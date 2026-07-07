#!/usr/bin/env python3
"""Build ~/Downloads/Neil_Job_Tracker.xlsx from the hub's career data.

Reads data/career-applications.json (+ career-networking.json for contact/
outreach info) and writes a single-sheet, one-line-per-field tracker ordered
most-important-info-first: date applied, company, job title, job type,
company type, status (+date), outcome (+date), salary, contact info, then
supporting links/notes, with location pushed to the far right column.
Re-run any time; it always overwrites the same file. Run automatically as the
last step of /jobs (see ~/.claude/commands/jobs.md).
"""
import json
import os
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HUB = os.path.expanduser('~/Claude/neil-ai-hub')
OUT = os.path.expanduser('~/Downloads/Neil_Job_Tracker.xlsx')

STATUS_ORDER = [
    'To Apply', 'Applied', 'Received Response', 'Did Not Apply',
    'Followed Up', 'Interview Scheduled', 'Interview Completed', 'Offer', 'Rejected',
]
STATUS_COLOR = {
    'To Apply': 'F97316', 'Applied': '2563EB', 'Received Response': '0EA5E9',
    'Did Not Apply': '64748B', 'Followed Up': '06B6D4', 'Interview Scheduled': 'A855F7',
    'Interview Completed': 'F59E0B', 'Offer': '10B981', 'Rejected': '94A3B8',
}
OUTCOME_COLOR = {'Rejected': 'EF4444', 'Follow Up': '10B981', 'Will Be In Touch': 'F59E0B'}

# (header, width) in the exact order Neil asked for: most important first,
# location pushed to the far right since it's the least decision-relevant field.
COLUMNS = [
    ('Date Applied', 12), ('Company', 26), ('Job Title', 40), ('Job Type', 24),
    ('Company Type', 30), ('Status', 17), ('Status Date', 12),
    ('Outcome', 20), ('Outcome Date', 12), ('Salary', 16),
    ('Contact Name', 16), ('Contact LinkedIn', 30),
    ('Outreach Status', 14), ('Outreach Date', 12),
    ('Date Posted', 12), ('Deadline', 12),
    ('Job Posting Link', 30), ('Portal / Login Link', 30),
    ('Why This Role', 46), ('Response Summary', 46), ('Location', 26),
]

# --- Company type classification -------------------------------------------------

KNOWN_COMPANY_TYPE = {
    'goldman sachs': 'Investment Bank',
    'natixis corporate & investment banking': 'Investment Bank',
    'td securities': 'Investment Bank',
    'oppenheimer & co.': 'Investment Bank / Broker-Dealer',
    'mufg': 'Bank',
    'santander us': 'Bank',
    'jpmorganchase': 'Bank',
    'bank of america': 'Bank',
    'u.s. bank': 'Bank',
    'citi': 'Bank',
    'smbc group': 'Bank',
    'bank of china usa': 'Bank',
    'bessemer trust': 'Private Bank / Trust Company',
    'cerberus capital management': 'Alternative Asset Manager / Private Equity',
    'vista equity partners': 'Alternative Asset Manager / Private Equity',
    'blackstone': 'Alternative Asset Manager / Private Equity',
    'blue owl capital': 'Alternative Asset Manager / Private Credit',
    'soros fund management': 'Hedge Fund',
    'brown advisory': 'Asset Manager / Wealth Management',
    'lord abbett': 'Asset Manager',
    'neuberger berman': 'Asset Manager',
    'schroders': 'Asset Manager',
    'willow wealth': 'Wealth Management Firm',
    'pwc': 'Consulting Firm (Big 4)',
    'talan': 'Consulting / Staffing Firm',
    'optiver': 'Proprietary Trading Firm',
    'tremendous': 'Fintech Company',
    'affirm': 'Fintech Company',
    'dtcc': 'Financial Market Infrastructure',
    'enstar group': 'Insurance / Reinsurance Group',
}

# Fallback keyword heuristics for companies not in KNOWN_COMPANY_TYPE (new
# postings going forward). Checked in order; first hit wins.
COMPANY_TYPE_KEYWORDS = [
    ('trust', 'Private Bank / Trust Company'),
    ('bank', 'Bank'),
    ('securities', 'Investment Bank'),
    ('capital management', 'Alternative Asset Manager / Private Equity'),
    ('equity partners', 'Alternative Asset Manager / Private Equity'),
    ('search', 'Recruiting / Staffing Firm'),
    ('partners', 'Asset Manager / Investment Firm'),
    ('group', 'Financial Services Firm'),
]

RECRUITER_CLIENT_HINTS = [
    ('hedge fund', 'hedge fund client'),
    ('prop trading firm', 'prop trading firm client'),
    ('prop trading', 'prop trading firm client'),
    ('family office', 'family office client'),
    ('private equity', 'private equity client'),
]


def split_company(raw):
    """Split 'Carisbrook Partners (recruiter, top global hedge fund)' into
    ('Carisbrook Partners', 'recruiter, top global hedge fund')."""
    m = re.match(r'^(.*?)\s*\(([^()]+)\)\s*$', raw or '')
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return (raw or '').strip(), ''


def classify_company_type(raw_company):
    name, hint = split_company(raw_company)
    hint_l = hint.lower()
    if 'recruiter' in hint_l:
        label = 'Recruiting / Staffing Firm'
        for keyword, suffix in RECRUITER_CLIENT_HINTS:
            if keyword in hint_l:
                return f'{label} — {suffix}'
        return label

    name_l = name.lower()
    if name_l in KNOWN_COMPANY_TYPE:
        return KNOWN_COMPANY_TYPE[name_l]
    for keyword, label in COMPANY_TYPE_KEYWORDS:
        if keyword in name_l:
            return label
    return 'Financial Services Firm'


# --- Job type classification -----------------------------------------------------

JOB_TYPE_RULES = [
    ('corporate action', 'Corporate Actions'),
    ('custody', 'Custody Operations'),
    ('settlement', 'Settlements'),
    ('reconciliation', 'Reconciliation'),
    ('trade finance', 'Trade Finance'),
    ('trade support', 'Trade Support'),
    ('trading services', 'Trade Support'),
    ('middle office', 'Middle Office'),
    ('due diligence', 'Due Diligence'),
    ('payments', 'Payments Operations'),
    ('treasury', 'Treasury Operations'),
    ('cash management', 'Treasury Operations'),
    ('capital operations', 'Treasury Operations'),
    ('valuation', 'Pricing & Valuation'),
    ('pricing', 'Pricing & Valuation'),
    ('controller', 'Controllers / Accounting'),
    ('accountant', 'Fund Accounting'),
    ('reporting', 'Reporting'),
    ('clo', 'Structured Products Operations'),
    ('structured finance', 'Structured Finance'),
    ('strategy', 'Strategy & Business Development'),
    ('business development', 'Strategy & Business Development'),
    ('deal', 'Deal / Valuation Advisory'),
    ('gcib', 'Corporate & Investment Banking'),
    ('investment banking', 'Corporate & Investment Banking'),
    ('wealth', 'Private Wealth / Client Service'),
    ('private client', 'Private Wealth / Client Service'),
    ('client', 'Client Service'),
    ('investor services', 'Client Service'),
    ('investor support', 'Client Service'),
    ('alternative investments', 'Alternative Investments Support'),
    ('investment operations', 'Investment Operations'),
    ('finance & operations', 'Finance & Operations'),
]


def classify_job_type(job_title):
    title_l = (job_title or '').lower()
    for keyword, label in JOB_TYPE_RULES:
        if keyword in title_l:
            return label
    return 'Operations'


def extract_location(job_title):
    """Split 'Custody Associate (New York, NY)' into ('Custody Associate', 'New York, NY')."""
    m = re.match(r'^(.*?)\s*\(([^()]+)\)\s*$', job_title or '')
    if m:
        return m.group(1).strip().rstrip(','), m.group(2).strip()
    return (job_title or '').strip(), ''


def tint(hex_color, amount=0.82):
    """Blend a hex color toward white so text stays readable (mimics the app's badge look)."""
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = round(r + (255 - r) * amount)
    g = round(g + (255 - g) * amount)
    b = round(b + (255 - b) * amount)
    return f'{r:02X}{g:02X}{b:02X}'


def one_line(text, limit=180):
    if not text:
        return ''
    flat = ' '.join(str(text).split())
    return flat if len(flat) <= limit else flat[:limit - 1] + '…'


def normco(name):
    return re.sub(r'[^a-z0-9]', '', re.sub(r'\(.*?\)', '', (name or '').lower()))


def batch_of(a):
    return a.get('batch') or (a.get('_created') or '')[:10]


def date_only(iso_str):
    if not iso_str:
        return ''
    return str(iso_str)[:10]


def status_date(a):
    history = a.get('statusHistory') or []
    if history:
        return date_only(history[-1].get('at'))
    return date_only(a.get('noApplyAt')) or ''


def outcome_and_date(a):
    if a.get('gmailOutcome'):
        return a['gmailOutcome'], date_only(a.get('gmailCheckedAt'))
    if a.get('noApplyReason'):
        return one_line(a['noApplyReason'], 80), date_only(a.get('noApplyAt'))
    return '', ''


def load(path):
    with open(path) as f:
        return json.load(f).get('data', [])


def match_contact(app, contacts):
    link = app.get('link')
    company = normco(app.get('company'))
    for c in contacts:
        if link and c.get('appLink') and c['appLink'] == link:
            return c
    for c in contacts:
        if normco(c.get('company')) == company:
            return c
    return None


def main():
    apps = load(os.path.join(HUB, 'data', 'career-applications.json'))
    contacts = load(os.path.join(HUB, 'data', 'career-networking.json'))

    def sort_key(a):
        status = a.get('status', 'To Apply')
        rank = STATUS_ORDER.index(status) if status in STATUS_ORDER else len(STATUS_ORDER)
        date = a.get('dateApplied') or batch_of(a) or ''
        return (rank, -_date_ordinal(date))

    rows = sorted(apps, key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Job Tracker'

    n_cols = len(COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    now = datetime.now(timezone.utc).strftime('%b %d, %Y %I:%M %p UTC')
    title_cell.value = f"Neil's Job Application Tracker — updated {now} — {len(rows)} total"
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1F2937')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 26

    header_row = 2
    for i, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 20

    STATUS_COL = 6
    OUTCOME_COL = 8
    LINK_COLS = {17: True, 18: True}  # Job Posting Link, Portal / Login Link
    CONTACT_LINK_COL = 12

    for r, a in enumerate(rows, start=header_row + 1):
        status = a.get('status', 'To Apply')
        outcome, outcome_dt = outcome_and_date(a)
        contact = match_contact(a, contacts) or {}
        job_title, location = extract_location(a.get('jobTitle', ''))
        company_name, _ = split_company(a.get('company', ''))

        values = [
            a.get('dateApplied', ''),
            company_name,
            one_line(job_title),
            classify_job_type(job_title),
            classify_company_type(a.get('company', '')),
            status,
            status_date(a),
            outcome,
            outcome_dt,
            one_line(a.get('salary', '')),
            contact.get('contactName', ''),
            contact.get('contactUrl', ''),
            contact.get('status', ''),
            date_only(contact.get('contactedAt', '')),
            batch_of(a),
            a.get('deadline', ''),
            a.get('link', ''),
            a.get('portalLink', ''),
            one_line(a.get('notes', '')),
            one_line(a.get('gmailSummary', '')),
            location,
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        status_cell = ws.cell(row=r, column=STATUS_COL)
        status_hex = STATUS_COLOR.get(status, '94A3B8')
        status_cell.fill = PatternFill('solid', fgColor=tint(status_hex))
        status_cell.font = Font(color=status_hex, bold=True)

        if outcome:
            outcome_cell = ws.cell(row=r, column=OUTCOME_COL)
            outcome_hex = OUTCOME_COLOR.get(outcome, '94A3B8')
            outcome_cell.fill = PatternFill('solid', fgColor=tint(outcome_hex))
            outcome_cell.font = Font(color=outcome_hex, bold=True)

        for col_idx in list(LINK_COLS) + [CONTACT_LINK_COL]:
            cell = ws.cell(row=r, column=col_idx)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color='2563EB', underline='single')

    last_row = header_row + len(rows)
    if rows:
        table_ref = f'A{header_row}:{last_col_letter}{last_row}'
        table = Table(displayName='JobTracker', ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False,
        )
        ws.add_table(table)

    ws.freeze_panes = f'A{header_row + 1}'

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'Wrote {len(rows)} rows to {OUT}')


def _date_ordinal(date_str):
    """Sort-friendly integer for YYYY-MM-DD strings; blank/unparseable sorts as 0 (oldest)."""
    try:
        return int((date_str or '0000-00-00').replace('-', ''))
    except ValueError:
        return 0


if __name__ == '__main__':
    main()
